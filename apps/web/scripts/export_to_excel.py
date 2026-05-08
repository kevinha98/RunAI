"""
Export strava-stats.json to a professional Excel workbook.
Sheets: Summary | Run Log | Monthly Trends | Yearly Trends | All Activities
"""
import json
import os
from collections import defaultdict
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ── paths ────────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_PATH = os.path.join(BASE, "data", "strava-stats.json")
OUT_PATH  = os.path.join(BASE, "data", "RunAI_Strava_Data_v2.xlsx")

# ── helpers ──────────────────────────────────────────────────────────────────
def load():
    with open(JSON_PATH, encoding="utf-8") as f:
        return json.load(f)

def m_to_km(m):  return round(m / 1000, 2) if m else 0
def s_to_min(s): return round(s / 60, 1)   if s else 0
def pace(dist_m, secs):
    """sec/km → mm:ss string"""
    if not dist_m or not secs: return ""
    spk = secs / (dist_m / 1000)
    mins = int(spk // 60)
    secs_ = int(spk % 60)
    return f"{mins}:{secs_:02d}"

def parse_dt(s):
    if not s: return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None

# ── style helpers ─────────────────────────────────────────────────────────────
DARK_ORANGE   = "FFE05C00"   # header bg
WHITE         = "FFFFFFFF"
LIGHT_GRAY    = "FFF2F2F2"
MID_GRAY      = "FFD9D9D9"
ACCENT        = "FF2563EB"   # bright blue for section headers
FONT_NAME     = "Calibri"

thin = Side(border_style="thin", color="FFB0B0B0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="FF333333"):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def apply_table_header(ws, row, cols, labels, col_widths=None):
    for i, label in enumerate(labels, 1):
        c = ws.cell(row=row, column=cols[0] + i - 1, value=label)
        c.font  = hdr_font()
        c.fill  = fill(DARK_ORANGE)
        c.alignment = center()
        c.border = BORDER
    if col_widths:
        for i, w in enumerate(col_widths, cols[0]):
            ws.column_dimensions[get_column_letter(i)].width = w

def style_data_row(ws, row_idx, start_col, end_col, alt=False):
    bg = LIGHT_GRAY if alt else WHITE
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row_idx, column=col)
        c.font = body_font()
        c.fill = fill(bg)
        c.border = BORDER
        if c.alignment is None or not c.alignment.horizontal:
            c.alignment = left()

def section_header(ws, row, col, text, span=1, bg=ACCENT):
    c = ws.cell(row=row, column=col, value=text)
    c.font = hdr_font(size=12)
    c.fill = fill(bg)
    c.alignment = left()
    c.border = BORDER
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
def build_summary(wb, data):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    athlete  = data.get("athlete", {})
    stats    = data.get("stravaStats", {})
    computed = data.get("computed", {})
    last_sync = parse_dt(data.get("lastSync"))

    # Title row
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "RunAI — Strava Performance Report"
    t.font  = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
    t.fill  = fill(DARK_ORANGE)
    t.alignment = center()

    # Sub-title
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"Athlete: {athlete.get('firstname','')} {athlete.get('lastname','')}  |  {athlete.get('city','')}, {athlete.get('country','')}  |  Last sync: {last_sync.strftime('%d %b %Y %H:%M') if last_sync else 'N/A'} UTC"
    sub.font = body_font(size=10, color="FF666666")
    sub.fill = fill(MID_GRAY)
    sub.alignment = center()

    # ── Section A: This Week ──────────────────────────────────────────────────
    section_header(ws, 4, 1, "This Week", span=6)
    labels_week = ["Weekly Km", "Weekly Runs", "Avg Pace (min/km)", "Longest Run 30d (km)", "YTD Runs", "YTD Km"]
    vals_week = [
        computed.get("weeklyKm", 0),
        computed.get("weeklyRuns", 0),
        "",   # formula below
        computed.get("longestRunKm", 0),
        stats.get("ytd_run_totals", {}).get("count", 0),
        m_to_km(stats.get("ytd_run_totals", {}).get("distance", 0)),
    ]
    apply_table_header(ws, 5, list(range(1,7)), labels_week, [16, 14, 20, 22, 12, 12])
    for i, (lbl, val) in enumerate(zip(labels_week, vals_week), 1):
        c = ws.cell(row=6, column=i, value=val)
        c.font = Font(name=FONT_NAME, size=13, bold=True, color=ACCENT.replace("FF",""))
        c.fill = fill(WHITE)
        c.border = BORDER
        c.alignment = center()
        if isinstance(val, float):
            c.number_format = '#,##0.0'

    # pace from computed
    avg_pace_spk = computed.get("avgPaceSecPerKm", 0)
    if avg_pace_spk:
        mins = int(avg_pace_spk // 60)
        secs_ = int(avg_pace_spk % 60)
        ws.cell(row=6, column=3, value=f"{mins}:{secs_:02d}")

    # ── Section B: Lifetime Stats ─────────────────────────────────────────────
    section_header(ws, 8, 1, "Lifetime Stats (All Runs)", span=6)
    all_run = stats.get("all_run_totals", {})
    labels_lt = ["Total Runs", "Total Distance (km)", "Total Moving Time (h)", "Total Elevation (m)", "Avg Distance/Run (km)", "Avg Moving Time/Run (min)"]
    apply_table_header(ws, 9, list(range(1,7)), labels_lt)
    total_runs = all_run.get("count", 0)
    total_dist_km = m_to_km(all_run.get("distance", 0))
    total_time_h  = round(all_run.get("moving_time", 0) / 3600, 1)
    total_elev    = round(all_run.get("elevation_gain", 0), 0)

    lifetime_vals = [total_runs, total_dist_km, total_time_h, total_elev,
                     f"=B10/{max(total_runs,1)}", f"=C10*60/{max(total_runs,1)}"]
    for i, val in enumerate(lifetime_vals, 1):
        c = ws.cell(row=10, column=i, value=val)
        c.font = body_font(size=11, bold=True)
        c.fill = fill(LIGHT_GRAY)
        c.border = BORDER
        c.alignment = center()
        if isinstance(val, (int, float)):
            c.number_format = '#,##0.0'

    # ── Section C: Recent Run Totals (28d) ───────────────────────────────────
    section_header(ws, 12, 1, "Recent 28-day Totals", span=6)
    rec = stats.get("recent_run_totals", {})
    labels_rec = ["Runs", "Distance (km)", "Moving Time (h)", "Elevation (m)", "Achievements", ""]
    apply_table_header(ws, 13, list(range(1,7)), labels_rec)
    rec_vals = [
        rec.get("count", 0),
        m_to_km(rec.get("distance", 0)),
        round(rec.get("moving_time", 0) / 3600, 1),
        round(rec.get("elevation_gain", 0), 0),
        rec.get("achievement_count", 0),
        "",
    ]
    for i, val in enumerate(rec_vals, 1):
        c = ws.cell(row=14, column=i, value=val)
        c.font = body_font(size=11)
        c.fill = fill(WHITE)
        c.border = BORDER
        c.alignment = center()
        if isinstance(val, (int, float)) and val:
            c.number_format = '#,##0.0'

    # col widths already set above; freeze top rows
    ws.freeze_panes = "A3"

# ── Sheet 2: Run Log ──────────────────────────────────────────────────────────
def build_run_log(wb, data):
    ws = wb.create_sheet("Run Log")
    ws.sheet_view.showGridLines = False

    runs = [a for a in data.get("recentActivities", []) if a.get("type") == "Run"]

    # Title
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = f"Run Log — {len(runs)} most recent runs"
    t.font  = Font(name=FONT_NAME, size=14, bold=True, color=WHITE)
    t.fill  = fill(DARK_ORANGE)
    t.alignment = center()

    HEADERS = ["#", "Date", "Name", "Distance (km)", "Duration (min)", "Pace (min/km)",
               "Elevation (m)", "Avg Speed (km/h)", "Avg Cadence", "PRs", "Kudos"]
    WIDTHS   = [5, 14, 32, 14, 15, 15, 13, 16, 14, 8, 8]

    apply_table_header(ws, 2, list(range(1, len(HEADERS)+1)), HEADERS, WIDTHS)

    for idx, run in enumerate(runs, 1):
        r = idx + 2
        alt = idx % 2 == 0
        dist_m  = run.get("distance", 0)
        mov_s   = run.get("moving_time", 0)
        elev    = run.get("total_elevation_gain", 0)
        avg_spd = run.get("average_speed", 0)  # m/s
        cad     = run.get("average_cadence", None)
        prs     = run.get("pr_count", 0)
        kudos   = run.get("kudos_count", 0)
        dt      = parse_dt(run.get("start_date_local") or run.get("start_date"))

        row_vals = [
            idx,
            dt,
            run.get("name", ""),
            round(dist_m / 1000, 2) if dist_m else 0,
            round(mov_s / 60, 1)    if mov_s else 0,
            pace(dist_m, mov_s),
            round(elev, 0),
            round(avg_spd * 3.6, 2) if avg_spd else 0,
            round(cad, 1) if cad else "",
            prs,
            kudos,
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            bg = LIGHT_GRAY if alt else WHITE
            c.font   = body_font()
            c.fill   = fill(bg)
            c.border = BORDER
            c.alignment = center() if col != 3 else left()

            # date format
            if col == 2 and isinstance(val, datetime):
                c.number_format = "DD MMM YYYY"
            # km / min — 1 decimal
            if col in (4, 5, 8):
                c.number_format = "#,##0.0"

    # Totals row
    n = len(runs)
    tr = n + 3
    ws.cell(row=tr, column=1, value="TOTALS").font = hdr_font(color="FF333333")
    ws.cell(row=tr, column=1).fill = fill(MID_GRAY)
    ws.cell(row=tr, column=1).border = BORDER
    ws.cell(row=tr, column=1).alignment = center()

    for col, formula in [
        (4, f"=SUM(D3:D{tr-1})"),
        (5, f"=SUM(E3:E{tr-1})"),
        (7, f"=SUM(G3:G{tr-1})"),
        (10, f"=SUM(J3:J{tr-1})"),
        (11, f"=SUM(K3:K{tr-1})"),
    ]:
        c = ws.cell(row=tr, column=col, value=formula)
        c.font = hdr_font(color="FF000000")
        c.fill = fill(MID_GRAY)
        c.border = BORDER
        c.alignment = center()
        c.number_format = "#,##0.0"

    for col in range(2, 12):
        c = ws.cell(row=tr, column=col)
        if c.value is None:
            c.fill = fill(MID_GRAY)
            c.border = BORDER

    # Avg pace row
    ar = tr + 1
    ws.cell(row=ar, column=1, value="AVG PACE").font = hdr_font(color="FF333333")
    ws.cell(row=ar, column=1).fill = fill(LIGHT_GRAY)
    ws.cell(row=ar, column=1).border = BORDER
    ws.cell(row=ar, column=1).alignment = center()
    c6 = ws.cell(row=ar, column=6, value=f'=IFERROR(TEXT(AVERAGE(F3:F{tr-1}),"mm:ss"),"")')
    c6.font = hdr_font(color="FF000000")
    c6.fill = fill(LIGHT_GRAY)
    c6.border = BORDER
    c6.alignment = center()
    for col in range(2, 12):
        c = ws.cell(row=ar, column=col)
        if c.value is None:
            c.fill = fill(LIGHT_GRAY)
            c.border = BORDER

    ws.freeze_panes = "A3"

# ── Sheet 3: All Activities ────────────────────────────────────────────────────
def build_all_activities(wb, data):
    ws = wb.create_sheet("All Activities")
    ws.sheet_view.showGridLines = False

    acts = data.get("recentActivities", [])

    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = f"All Activities — {len(acts)} records"
    t.font  = Font(name=FONT_NAME, size=14, bold=True, color=WHITE)
    t.fill  = fill(DARK_ORANGE)
    t.alignment = center()

    HEADERS = ["#", "Date", "Type", "Name", "Distance (km)", "Duration (min)",
               "Elevation (m)", "Avg Speed (km/h)", "PRs", "Kudos"]
    WIDTHS   = [5, 14, 12, 36, 14, 15, 13, 16, 8, 8]
    apply_table_header(ws, 2, list(range(1, len(HEADERS)+1)), HEADERS, WIDTHS)

    for idx, act in enumerate(acts, 1):
        r   = idx + 2
        alt = idx % 2 == 0
        dist_m  = act.get("distance", 0)
        mov_s   = act.get("moving_time", 0)
        avg_spd = act.get("average_speed", 0)
        dt      = parse_dt(act.get("start_date_local") or act.get("start_date"))

        row_vals = [
            idx,
            dt,
            act.get("sport_type") or act.get("type", ""),
            act.get("name", ""),
            round(dist_m / 1000, 2) if dist_m else 0,
            round(mov_s / 60, 1)    if mov_s else 0,
            round(act.get("total_elevation_gain", 0), 0),
            round(avg_spd * 3.6, 2) if avg_spd else 0,
            act.get("pr_count", 0),
            act.get("kudos_count", 0),
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            bg = LIGHT_GRAY if alt else WHITE
            c.font   = body_font()
            c.fill   = fill(bg)
            c.border = BORDER
            c.alignment = center() if col not in (3, 4) else left()
            if col == 2 and isinstance(val, datetime):
                c.number_format = "DD MMM YYYY"
            if col in (5, 6, 8):
                c.number_format = "#,##0.0"

    ws.freeze_panes = "A3"

# ── Sheet 4: Yearly Stats ─────────────────────────────────────────────────────
def build_yearly(wb, data):
    ws = wb.create_sheet("Yearly Stats")
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Yearly Stats — YTD vs Lifetime"
    t.font  = Font(name=FONT_NAME, size=14, bold=True, color=WHITE)
    t.fill  = fill(DARK_ORANGE)
    t.alignment = center()

    stats = data.get("stravaStats", {})
    ytd   = stats.get("ytd_run_totals", {})
    all_  = stats.get("all_run_totals", {})

    HEADERS = ["Metric", "YTD", "Lifetime", "YTD %", "", ""]
    WIDTHS   = [28, 16, 16, 14, 14, 14]
    apply_table_header(ws, 2, list(range(1,7)), HEADERS, WIDTHS)

    rows = [
        ("Runs",              ytd.get("count",0),                           all_.get("count",0)),
        ("Distance (km)",     m_to_km(ytd.get("distance",0)),               m_to_km(all_.get("distance",0))),
        ("Moving Time (h)",   round(ytd.get("moving_time",0)/3600,1),       round(all_.get("moving_time",0)/3600,1)),
        ("Elevation Gain (m)",round(ytd.get("elevation_gain",0),0),         round(all_.get("elevation_gain",0),0)),
    ]

    for i, (metric, ytd_val, all_val) in enumerate(rows, 1):
        r   = i + 2
        alt = i % 2 == 0
        bg  = LIGHT_GRAY if alt else WHITE

        c1 = ws.cell(row=r, column=1, value=metric)
        c1.font = body_font(bold=True); c1.fill = fill(bg); c1.border = BORDER; c1.alignment = left()

        c2 = ws.cell(row=r, column=2, value=ytd_val)
        c2.font = body_font(); c2.fill = fill(bg); c2.border = BORDER; c2.alignment = center()
        c2.number_format = "#,##0.0"

        c3 = ws.cell(row=r, column=3, value=all_val)
        c3.font = body_font(); c3.fill = fill(bg); c3.border = BORDER; c3.alignment = center()
        c3.number_format = "#,##0.0"

        c4 = ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/C{r},0)")
        c4.font = body_font(); c4.fill = fill(bg); c4.border = BORDER; c4.alignment = center()
        c4.number_format = "0.0%"

        for col in (5, 6):
            cx = ws.cell(row=r, column=col)
            cx.fill = fill(bg); cx.border = BORDER

    ws.freeze_panes = "A3"

# ── Aggregate run data by month and year ──────────────────────────────────────
def aggregate_runs(data):
    runs = [a for a in data.get("recentActivities", [])
            if a.get("type") == "Run" or a.get("sport_type") == "Run"]

    monthly = defaultdict(lambda: {"runs": 0, "dist_m": 0, "time_s": 0, "elev": 0, "pr": 0})
    yearly  = defaultdict(lambda: {"runs": 0, "dist_m": 0, "time_s": 0, "elev": 0, "pr": 0})

    for r in runs:
        dt = parse_dt(r.get("start_date_local") or r.get("start_date"))
        if not dt:
            continue
        mk = f"{dt.year}-{dt.month:02d}"
        yk = str(dt.year)
        for d in (monthly[mk], yearly[yk]):
            d["runs"]   += 1
            d["dist_m"] += r.get("distance", 0)
            d["time_s"] += r.get("moving_time", 0)
            d["elev"]   += r.get("total_elevation_gain", 0)
            d["pr"]     += r.get("pr_count", 0)

    # Sort keys chronologically
    sorted_months = sorted(monthly.keys())
    sorted_years  = sorted(yearly.keys())
    return monthly, yearly, sorted_months, sorted_years

def pace_str(dist_m, time_s):
    if not dist_m or not time_s: return ""
    spk = time_s / (dist_m / 1000)
    return f"{int(spk//60)}:{int(spk%60):02d}"

# ── Sheet: Monthly Trends ─────────────────────────────────────────────────────
def build_monthly_trends(wb, data):
    ws = wb.create_sheet("Monthly Trends")
    ws.sheet_view.showGridLines = False

    monthly, _, sorted_months, _ = aggregate_runs(data)

    # Title
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"Monthly Trends — {len(sorted_months)} months of running data"
    t.font  = Font(name=FONT_NAME, size=14, bold=True, color=WHITE)
    t.fill  = fill(DARK_ORANGE)
    t.alignment = center()

    HEADERS = ["Month", "Runs", "Distance (km)", "Moving Time (h)",
               "Elevation (m)", "Avg Pace (min/km)", "PRs", "Avg Dist/Run (km)"]
    WIDTHS  = [14, 8, 16, 16, 14, 18, 8, 18]
    apply_table_header(ws, 2, list(range(1, 9)), HEADERS, WIDTHS)

    for i, mk in enumerate(sorted_months, 1):
        r   = i + 2
        alt = i % 2 == 0
        d   = monthly[mk]
        dist_km = round(d["dist_m"] / 1000, 1)
        time_h  = round(d["time_s"] / 3600, 1)
        elev    = round(d["elev"], 0)
        p       = pace_str(d["dist_m"], d["time_s"])
        avg_d   = round(dist_km / d["runs"], 1) if d["runs"] else 0

        vals = [mk, d["runs"], dist_km, time_h, elev, p, d["pr"], avg_d]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            bg = LIGHT_GRAY if alt else WHITE
            c.font = body_font(); c.fill = fill(bg); c.border = BORDER
            c.alignment = center() if col != 1 else left()
            if col in (3, 4, 8): c.number_format = "#,##0.0"

    data_rows = len(sorted_months)
    last_data = data_rows + 2

    # ── Totals row ────────────────────────────────────────────────────────────
    tr = last_data + 1
    labels = ["TOTAL", f"=SUM(B3:B{last_data})", f"=SUM(C3:C{last_data})",
              f"=SUM(D3:D{last_data})", f"=SUM(E3:E{last_data})", "", f"=SUM(G3:G{last_data})", ""]
    for col, val in enumerate(labels, 1):
        c = ws.cell(row=tr, column=col, value=val)
        c.font = hdr_font(color="FF000000"); c.fill = fill(MID_GRAY)
        c.border = BORDER; c.alignment = center()
        if col in (3, 4, 8): c.number_format = "#,##0.0"
    ws.cell(row=tr, column=1).alignment = left()

    # ── Chart 1: km per month (bar) ───────────────────────────────────────────
    months_ref = Reference(ws, min_col=1, min_row=3, max_row=last_data)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Distance per Month (km)"
    chart1.y_axis.title = "Distance (km)"
    chart1.y_axis.numFmt = '#,##0'
    chart1.x_axis.title = "Month"
    chart1.x_axis.tickLblSkip = max(1, len(sorted_months) // 12)
    chart1.style = 10
    chart1.width  = 36
    chart1.height = 14

    km_data = Reference(ws, min_col=3, min_row=2, max_row=last_data)  # row 2 = header "Distance (km)"
    chart1.add_data(km_data, titles_from_data=True)
    chart1.set_categories(months_ref)
    chart1.series[0].graphicalProperties.solidFill = "E05C00"
    chart1.series[0].graphicalProperties.line.solidFill = "E05C00"
    ws.add_chart(chart1, f"A{tr + 2}")

    # ── Chart 2: runs per month (line) ────────────────────────────────────────
    chart2 = LineChart()
    chart2.title = "Runs per Month"
    chart2.y_axis.title = "Number of Runs"
    chart2.y_axis.numFmt = '0'
    chart2.x_axis.title = "Month"
    chart2.x_axis.tickLblSkip = max(1, len(sorted_months) // 12)
    chart2.style = 10
    chart2.width  = 36
    chart2.height = 14

    runs_data = Reference(ws, min_col=2, min_row=2, max_row=last_data)  # row 2 = header "Runs"
    chart2.add_data(runs_data, titles_from_data=True)
    chart2.set_categories(months_ref)
    chart2.series[0].graphicalProperties.line.solidFill = "2563EB"
    chart2.series[0].graphicalProperties.line.width = 20000
    ws.add_chart(chart2, f"L{tr + 2}")

    ws.freeze_panes = "A3"

# ── Sheet: Yearly Trends ──────────────────────────────────────────────────────
def build_yearly_trends(wb, data):
    ws = wb.create_sheet("Yearly Trends")
    ws.sheet_view.showGridLines = False

    _, yearly, _, sorted_years = aggregate_runs(data)

    # Title
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"Yearly Trends — {len(sorted_years)} years of running data"
    t.font  = Font(name=FONT_NAME, size=14, bold=True, color=WHITE)
    t.fill  = fill(DARK_ORANGE)
    t.alignment = center()

    HEADERS = ["Year", "Runs", "Distance (km)", "Moving Time (h)",
               "Elevation (m)", "Avg Pace (min/km)", "PRs", "Avg Dist/Run (km)"]
    WIDTHS  = [10, 8, 16, 16, 14, 18, 8, 18]
    apply_table_header(ws, 2, list(range(1, 9)), HEADERS, WIDTHS)

    for i, yk in enumerate(sorted_years, 1):
        r   = i + 2
        alt = i % 2 == 0
        d   = yearly[yk]
        dist_km = round(d["dist_m"] / 1000, 1)
        time_h  = round(d["time_s"] / 3600, 1)
        elev    = round(d["elev"], 0)
        p       = pace_str(d["dist_m"], d["time_s"])
        avg_d   = round(dist_km / d["runs"], 1) if d["runs"] else 0

        vals = [yk, d["runs"], dist_km, time_h, elev, p, d["pr"], avg_d]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            bg = LIGHT_GRAY if alt else WHITE
            c.font = body_font(bold=(col==1)); c.fill = fill(bg); c.border = BORDER
            c.alignment = center() if col != 1 else left()
            if col in (3, 4, 8): c.number_format = "#,##0.0"

    data_rows = len(sorted_years)
    last_data = data_rows + 2

    # Totals
    tr = last_data + 1
    labels = ["TOTAL", f"=SUM(B3:B{last_data})", f"=SUM(C3:C{last_data})",
              f"=SUM(D3:D{last_data})", f"=SUM(E3:E{last_data})", "", f"=SUM(G3:G{last_data})", ""]
    for col, val in enumerate(labels, 1):
        c = ws.cell(row=tr, column=col, value=val)
        c.font = hdr_font(color="FF000000"); c.fill = fill(MID_GRAY)
        c.border = BORDER; c.alignment = center()
        if col in (3, 4, 8): c.number_format = "#,##0.0"
    ws.cell(row=tr, column=1).alignment = left()

    years_ref = Reference(ws, min_col=1, min_row=3, max_row=last_data)

    # ── Chart 1: km per year (bar) ────────────────────────────────────────────
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Distance per Year (km)"
    chart1.y_axis.title = "Distance (km)"
    chart1.y_axis.numFmt = '#,##0'
    chart1.x_axis.title = "Year"
    chart1.style = 10
    chart1.width  = 22
    chart1.height = 14

    km_data = Reference(ws, min_col=3, min_row=2, max_row=last_data)  # row 2 = header
    chart1.add_data(km_data, titles_from_data=True)
    chart1.set_categories(years_ref)
    chart1.series[0].graphicalProperties.solidFill = "E05C00"
    chart1.series[0].graphicalProperties.line.solidFill = "E05C00"
    ws.add_chart(chart1, f"A{tr + 2}")

    # ── Chart 2: runs per year (bar) ──────────────────────────────────────────
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Runs per Year"
    chart2.y_axis.title = "Number of Runs"
    chart2.y_axis.numFmt = '0'
    chart2.x_axis.title = "Year"
    chart2.style = 10
    chart2.width  = 22
    chart2.height = 14

    runs_data = Reference(ws, min_col=2, min_row=2, max_row=last_data)  # row 2 = header
    chart2.add_data(runs_data, titles_from_data=True)
    chart2.set_categories(years_ref)
    chart2.series[0].graphicalProperties.solidFill = "2563EB"
    chart2.series[0].graphicalProperties.line.solidFill = "2563EB"
    ws.add_chart(chart2, f"K{tr + 2}")

    # ── Chart 3: elevation per year (bar) ─────────────────────────────────────
    chart3 = BarChart()
    chart3.type = "col"
    chart3.title = "Elevation Gain per Year (m)"
    chart3.y_axis.title = "Elevation (m)"
    chart3.y_axis.numFmt = '#,##0'
    chart3.x_axis.title = "Year"
    chart3.style = 10
    chart3.width  = 22
    chart3.height = 14

    elev_data = Reference(ws, min_col=5, min_row=2, max_row=last_data)  # row 2 = header
    chart3.add_data(elev_data, titles_from_data=True)
    chart3.set_categories(years_ref)
    chart3.series[0].graphicalProperties.solidFill = "16A34A"
    chart3.series[0].graphicalProperties.line.solidFill = "16A34A"
    ws.add_chart(chart3, f"A{tr + 24}")

    ws.freeze_panes = "A3"

# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print(f"Reading {JSON_PATH}")
    data = load()

    wb = Workbook()
    build_summary(wb, data)
    build_run_log(wb, data)
    build_monthly_trends(wb, data)
    build_yearly_trends(wb, data)
    build_all_activities(wb, data)

    wb.save(OUT_PATH)
    print(f"Saved → {OUT_PATH}")

if __name__ == "__main__":
    main()
